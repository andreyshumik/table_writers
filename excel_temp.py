import openpyxl

wb = openpyxl.load_workbook('temp.xlsx')
sheet = wb.sheetnames
sheet = wb.active

for i in range(0, sheet.max_row):
    for col in sheet.iter_cols(1, sheet.max_column):
        print(col[i].value, end="\t\t")
    print('')