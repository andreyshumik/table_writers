from abc import ABC, abstractmethod
import csv
import openpyxl

CellValue = str | int | float


class TableWriter(ABC):
    @abstractmethod
    def __init__(self, filename: str):
        pass

    @abstractmethod
    def write(self, data: list[list[CellValue]]):
        pass


class TableReader(ABC):
    @abstractmethod
    def __init__(self, filename: str):
        pass

    @abstractmethod
    def read(self, data: list[list[CellValue]]):
        pass


class CSVTableReader(TableReader):
    def __init__(self, filename: str):
        self._filename = filename

    def read(self) -> list[list[CellValue]]:
        r_file = open(self._filename, 'r')
        file_reader = csv.reader(r_file, delimiter=",")
        li = []
        count = 0
        for row in file_reader:
            print(row)
            li.append(row)
            count += 1
        return li


class CSVTableWriter(TableWriter):
    def __init__(self, filename: str):
        self._filename = filename

    def write(self, data: list[list[CellValue]]):
        fd = open(self._filename, 'w')
        headers, *contents = data
        csvwriter = csv.DictWriter(fd, fieldnames=headers)
        csvwriter.writeheader()

        for line in contents:
            csvwriter.writerow(dict(zip(headers, line)))



class XLSTableReader(TableReader):
    def __init__(self, filename: str):
        self._filename = filename

    def read(self) -> list[list[CellValue]]:
        wb = openpyxl.load_workbook(self._filename)
        sheet = wb.sheetnames
        sheet = wb.active

        for i in range(0, sheet.max_row):
            for col in sheet.iter_cols(1, sheet.max_column):
                print(col[i].value, end="\t\t")
            print('')


class HTMLTableWriter(TableWriter):
    def __init__(self, filename: str):
        self._filename = filename

    def _build_header(self, data: list[CellValue]) -> str:
        cells = [f'<th>{value}</th>' for value in data]
        return f'<thead><tr>{"".join(cells)}</tr></thead>'

    def _build_body(self, data: list[list[CellValue]]) -> str:
        lines = []
        for line in data:
            cells = [f'<td>{value}</td>' for value in line]
            lines.append(f'<tr>{"".join(cells)}</tr>')

        return f'<tbody>{"".join(lines)}</tbody>'

    def write(self, data: list[list[CellValue]]):
        fd = open(self._filename, 'w')
        fd.write('<html><head/><body><table>')
        fd.write(self._build_header(data[0]))
        fd.write(self._build_body(data[1:]))
        fd.write('</table></body></html>')
        fd.close()


if __name__ == "__main__":
    content = [
        ['Столбец1', 'Столбец2'],
        [0, 1],
        [2, 3],
    ]
    writer = HTMLTableWriter('output.html')
    writer.write(content)

    writer = CSVTableWriter('output.csv')
    writer.write(content)

    reader = CSVTableReader('output.csv')
    reader.read()

    reader1 = XLSTableReader('temp.xlsx')
    reader1.read()